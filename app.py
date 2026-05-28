from __future__ import annotations

import math
import os
import re
import sqlite3
import unicodedata
from datetime import datetime
from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from typing import Dict, Iterable, List, Optional, Tuple
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# Configuração geral
# ============================================================
MOEDA_Q = Decimal("0.01")
PRECO_Q = Decimal("0.0001")

APP_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(APP_DIR, "data")
DB_PATH = os.path.join(DATA_DIR, "precos_app.sqlite3")
os.makedirs(DATA_DIR, exist_ok=True)

CANON_PEDIDO = {
    "numero_pedido": ["Número pedido", "Numero pedido", "Pedido", "Nº Pedido", "N Pedido"],
    "nome_comprador": ["Nome Comprador", "Comprador"],
    "sku": ["SKU", "Código", "Codigo", "Cód. Produto", "Cod Produto"],
    "produto": ["Produto", "Descrição", "Descricao"],
    "unidade": ["Un", "Unidade"],
    "quantidade": ["Quantidade", "Qtd", "Qtde"],
    "valor_unitario": ["Valor Unitário", "Valor Unitario", "Preço", "Preco", "Valor Unit"],
    "valor_total": ["Valor Total", "Total Item", "Total do Item"],
    "total_pedido": ["Total Pedido", "Valor Pedido", "Total Geral"],
    "observacoes": ["Observações", "Observacoes", "Obs"],
}

CANON_PRECO = {
    "nome_lista": ["Nome lista", "Lista", "Tabela", "Nome da lista"],
    "sku": ["SKU", "Sku", "Código", "Codigo", "Cód. Produto", "Cod Produto"],
    "preco": ["Preço", "Preco", "Valor", "Valor Unitário", "Valor Unitario", "R$ Preço da lista", "Preço da lista", "Preco da lista"],
}


@dataclass
class ItemPedido:
    indice: int
    linha_excel: int
    valores_linha: List[object]
    sku: str
    quantidade: int
    preco_unitario: Decimal
    valor_total: Decimal


@dataclass
class ResultadoDivisao:
    sucesso: bool
    mensagem: str
    alocacao: List[Dict[int, int]]
    totais: List[Decimal]
    avisos: List[str]


# ============================================================
# Utilitários de leitura / normalização
# ============================================================
def normalizar_texto(valor: object) -> str:
    if valor is None:
        return ""
    texto = str(valor).strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = texto.lower()
    texto = re.sub(r"[^a-z0-9]+", "", texto)
    return texto


def normalizar_sku(valor: object) -> str:
    if valor is None:
        return ""
    texto = str(valor).strip().upper()
    texto = texto.replace("–", "-").replace("—", "-")
    texto = re.sub(r"\s+", "", texto)
    return texto


def variantes_sku(valor: object) -> List[str]:
    base = normalizar_sku(valor)
    if not base:
        return []
    variantes = {base}
    if "," in base:
        variantes.add(base.replace(",", "."))
    if "." in base:
        variantes.add(base.replace(".", ","))
    return sorted(variantes)


def registrar_preco(precos: Dict[str, Decimal], sku: object, preco: Decimal) -> None:
    for variante in variantes_sku(sku):
        precos[variante] = preco


def mapa_alias(canon: Dict[str, List[str]]) -> Dict[str, str]:
    saida = {}
    for chave, aliases in canon.items():
        for alias in aliases + [chave]:
            saida[normalizar_texto(alias)] = chave
    return saida


ALIAS_PEDIDO = mapa_alias(CANON_PEDIDO)
ALIAS_PRECO = mapa_alias(CANON_PRECO)


def para_decimal(valor: object, campo: str = "valor") -> Decimal:
    if valor is None or valor == "":
        raise ValueError(f"Campo {campo} vazio.")
    if isinstance(valor, Decimal):
        return valor
    texto = str(valor).strip()
    # Aceita formatos pt-BR simples: 1.234,56 ou 1234,56.
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        return Decimal(texto)
    except InvalidOperation as exc:
        raise ValueError(f"Não foi possível converter {campo}: {valor!r}") from exc


def fmt_moeda(valor: Decimal) -> str:
    valor = valor.quantize(MOEDA_Q, rounding=ROUND_HALF_UP)
    s = f"{valor:,.2f}"
    return "R$ " + s.replace(",", "X").replace(".", ",").replace("X", ".")


def localizar_cabecalho(ws, alias_map: Dict[str, str], obrigatorios: Iterable[str], max_linhas: int = 12) -> Tuple[int, Dict[str, int]]:
    melhor_linha = None
    melhor_mapa = {}
    melhor_score = -1
    obrigatorios = set(obrigatorios)

    for row in range(1, min(ws.max_row, max_linhas) + 1):
        mapa = {}
        for col in range(1, ws.max_column + 1):
            chave = alias_map.get(normalizar_texto(ws.cell(row, col).value))
            if chave and chave not in mapa:
                mapa[chave] = col
        score = len(obrigatorios.intersection(mapa.keys())) * 10 + len(mapa)
        if score > melhor_score:
            melhor_score = score
            melhor_linha = row
            melhor_mapa = mapa

    faltantes = obrigatorios.difference(melhor_mapa.keys())
    if faltantes:
        faltantes_txt = ", ".join(sorted(faltantes))
        raise ValueError(f"Cabeçalho não localizado ou colunas obrigatórias ausentes: {faltantes_txt}.")

    return melhor_linha or 1, melhor_mapa


def carregar_precos_pdf_bling(preco_bytes: bytes) -> Dict[str, Decimal]:
    """Extrai SKU e R$ Preço da lista do relatório PDF do Bling."""
    try:
        import pdfplumber
    except ImportError as exc:
        raise ImportError(
            "Para ler PDF do Bling, inclua pdfplumber no requirements.txt e reinstale as dependências."
        ) from exc

    precos: Dict[str, Decimal] = {}
    linhas_lidas = 0
    padrao_linha = re.compile(
        r"(?P<sku>[A-Z0-9][A-Z0-9\-\./,]*[A-Z0-9])\s+"
        r"(?:(?P<gtin>\d{8,14})\s+)?"
        r"(?P<preco_bling>\d{1,3}(?:\.\d{3})*,\d{2})\s+"
        r"(?P<preco_lista>\d{1,3}(?:\.\d{3})*,\d{2})\s*$"
    )

    with pdfplumber.open(BytesIO(preco_bytes)) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text(x_tolerance=1, y_tolerance=3) or ""
            for linha in texto.splitlines():
                linha = linha.strip()
                m = padrao_linha.search(linha)
                if not m:
                    continue
                sku = m.group("sku")
                preco_lista = para_decimal(m.group("preco_lista"), f"preço da lista do SKU {sku}").quantize(PRECO_Q, rounding=ROUND_HALF_UP)
                registrar_preco(precos, sku, preco_lista)
                linhas_lidas += 1

    if not precos:
        raise ValueError(
            "Não foi possível extrair preços do PDF. Confirme se é o relatório 'Listas de preços' do Bling com as colunas SKU e R$ Preço da lista."
        )
    return precos


def identificar_tipo_preco(nome_arquivo: str) -> str:
    nome = (nome_arquivo or "").lower()
    if nome.endswith(".pdf"):
        return "pdf"
    return "xlsx"


def listar_abas_e_listas_preco(preco_bytes: bytes) -> Tuple[List[str], List[str]]:
    wb = load_workbook(BytesIO(preco_bytes), data_only=True, read_only=True)
    nomes_listas = set()
    for ws in wb.worksheets:
        try:
            header_row, colmap = localizar_cabecalho(ws, ALIAS_PRECO, ["sku", "preco"])
        except Exception:
            continue
        col_lista = colmap.get("nome_lista")
        if col_lista:
            for row in range(header_row + 1, ws.max_row + 1):
                valor = ws.cell(row, col_lista).value
                if valor not in (None, ""):
                    nomes_listas.add(str(valor).strip())
    return wb.sheetnames, sorted(nomes_listas)


def carregar_precos(preco_bytes: bytes, aba: str, nome_lista: Optional[str]) -> Dict[str, Decimal]:
    wb = load_workbook(BytesIO(preco_bytes), data_only=True, read_only=True)
    ws = wb[aba]
    header_row, colmap = localizar_cabecalho(ws, ALIAS_PRECO, ["sku", "preco"])

    precos: Dict[str, Decimal] = {}
    for row in range(header_row + 1, ws.max_row + 1):
        sku_raw = ws.cell(row, colmap["sku"]).value
        preco_raw = ws.cell(row, colmap["preco"]).value
        if sku_raw in (None, "") and preco_raw in (None, ""):
            continue

        if nome_lista and colmap.get("nome_lista"):
            lista_atual = ws.cell(row, colmap["nome_lista"]).value
            if str(lista_atual).strip() != nome_lista:
                continue

        if sku_raw in (None, ""):
            continue
        sku = str(sku_raw).strip()
        if preco_raw in (None, ""):
            continue
        preco = para_decimal(preco_raw, f"preço do SKU {sku}").quantize(PRECO_Q, rounding=ROUND_HALF_UP)
        registrar_preco(precos, sku, preco)

    if not precos:
        raise ValueError("Nenhum preço válido foi encontrado na lista selecionada.")
    return precos


def carregar_itens_pedido(pedido_bytes: bytes, aba: str, precos: Dict[str, Decimal], skus_excluidos: Optional[set[str]] = None) -> Tuple[int, Dict[str, int], List[ItemPedido], List[Dict[str, object]]]:
    wb = load_workbook(BytesIO(pedido_bytes), data_only=False, read_only=False)
    ws = wb[aba]
    header_row, colmap = localizar_cabecalho(ws, ALIAS_PEDIDO, ["numero_pedido", "sku", "quantidade"])

    itens: List[ItemPedido] = []
    pendencias: List[Dict[str, object]] = []
    skus_excluidos = skus_excluidos or set()

    for row in range(header_row + 1, ws.max_row + 1):
        sku_raw = ws.cell(row, colmap["sku"]).value
        qtd_raw = ws.cell(row, colmap["quantidade"]).value

        if sku_raw in (None, "") and qtd_raw in (None, ""):
            continue
        if sku_raw in (None, ""):
            pendencias.append({"Linha": row, "SKU": "", "Problema": "Linha com quantidade, mas sem SKU."})
            continue

        sku = str(sku_raw).strip()
        sku_chave = normalizar_sku(sku)
        if sku_chave in skus_excluidos:
            continue
        if sku_chave not in precos:
            pendencias.append({"Linha": row, "SKU": sku, "Problema": "SKU não encontrado na lista de preços selecionada."})
            continue

        try:
            qtd_dec = para_decimal(qtd_raw, f"quantidade da linha {row}")
        except Exception as exc:
            pendencias.append({"Linha": row, "SKU": sku, "Problema": str(exc)})
            continue

        if qtd_dec <= 0:
            pendencias.append({"Linha": row, "SKU": sku, "Problema": "Quantidade menor ou igual a zero."})
            continue

        if qtd_dec != qtd_dec.to_integral_value():
            pendencias.append({"Linha": row, "SKU": sku, "Problema": "Quantidade fracionária; ajuste manualmente ou trate em pedido separado."})
            continue

        quantidade = int(qtd_dec)
        preco = precos[sku_chave]
        total = (preco * Decimal(quantidade)).quantize(MOEDA_Q, rounding=ROUND_HALF_UP)
        valores_linha = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        itens.append(
            ItemPedido(
                indice=len(itens),
                linha_excel=row,
                valores_linha=valores_linha,
                sku=sku,
                quantidade=quantidade,
                preco_unitario=preco,
                valor_total=total,
            )
        )

    return header_row, colmap, itens, pendencias


# ============================================================
# Banco de dados local: fontes, preços manuais e exclusões
# ============================================================
def conectar_banco() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def inicializar_banco() -> None:
    with conectar_banco() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS fontes_preco (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome_arquivo TEXT NOT NULL,
                tipo TEXT NOT NULL,
                conteudo BLOB NOT NULL,
                criado_em TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS precos (
                sku TEXT PRIMARY KEY,
                preco TEXT NOT NULL,
                origem TEXT NOT NULL,
                atualizado_em TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS skus_excluidos (
                sku TEXT PRIMARY KEY,
                motivo TEXT,
                atualizado_em TEXT NOT NULL
            )
            """
        )
        conn.commit()


def salvar_fonte_preco(nome_arquivo: str, tipo: str, conteudo: bytes) -> None:
    with conectar_banco() as conn:
        conn.execute(
            "INSERT INTO fontes_preco (nome_arquivo, tipo, conteudo, criado_em) VALUES (?, ?, ?, ?)",
            (nome_arquivo, tipo, conteudo, datetime.now().isoformat(timespec="seconds")),
        )
        conn.commit()


def salvar_precos_no_banco(precos: Dict[str, Decimal], origem: str) -> None:
    agora = datetime.now().isoformat(timespec="seconds")
    with conectar_banco() as conn:
        for sku, preco in precos.items():
            conn.execute(
                """
                INSERT INTO precos (sku, preco, origem, atualizado_em) VALUES (?, ?, ?, ?)
                ON CONFLICT(sku) DO UPDATE SET
                    preco=excluded.preco,
                    origem=excluded.origem,
                    atualizado_em=excluded.atualizado_em
                """,
                (normalizar_sku(sku), str(preco.quantize(PRECO_Q, rounding=ROUND_HALF_UP)), origem, agora),
            )
        conn.commit()


def carregar_precos_do_banco() -> Dict[str, Decimal]:
    inicializar_banco()
    precos: Dict[str, Decimal] = {}
    with conectar_banco() as conn:
        for sku, preco_txt in conn.execute("SELECT sku, preco FROM precos"):
            try:
                registrar_preco(precos, sku, Decimal(str(preco_txt)))
            except Exception:
                continue
    return precos


def carregar_skus_excluidos() -> set[str]:
    inicializar_banco()
    with conectar_banco() as conn:
        return {normalizar_sku(row[0]) for row in conn.execute("SELECT sku FROM skus_excluidos")}


def salvar_precos_manuais(linhas: List[Dict[str, object]]) -> int:
    agora = datetime.now().isoformat(timespec="seconds")
    total = 0
    with conectar_banco() as conn:
        for linha in linhas:
            sku = normalizar_sku(linha.get("SKU"))
            preco_raw = linha.get("Preço manual")
            excluir = bool(linha.get("Excluir do pedido"))
            if not sku:
                continue
            if excluir:
                conn.execute(
                    """
                    INSERT INTO skus_excluidos (sku, motivo, atualizado_em) VALUES (?, ?, ?)
                    ON CONFLICT(sku) DO UPDATE SET motivo=excluded.motivo, atualizado_em=excluded.atualizado_em
                    """,
                    (sku, "Excluído manualmente na tela de pendências", agora),
                )
                total += 1
                continue
            if preco_raw in (None, ""):
                continue
            preco = para_decimal(preco_raw, f"preço manual do SKU {sku}").quantize(PRECO_Q, rounding=ROUND_HALF_UP)
            conn.execute(
                """
                INSERT INTO precos (sku, preco, origem, atualizado_em) VALUES (?, ?, ?, ?)
                ON CONFLICT(sku) DO UPDATE SET
                    preco=excluded.preco, origem=excluded.origem, atualizado_em=excluded.atualizado_em
                """,
                (sku, str(preco), "manual", agora),
            )
            conn.execute("DELETE FROM skus_excluidos WHERE sku = ?", (sku,))
            total += 1
        conn.commit()
    return total


def listar_fontes_salvas() -> pd.DataFrame:
    inicializar_banco()
    with conectar_banco() as conn:
        rows = conn.execute(
            "SELECT id, nome_arquivo, tipo, criado_em FROM fontes_preco ORDER BY id DESC LIMIT 20"
        ).fetchall()
    return pd.DataFrame(rows, columns=["ID", "Arquivo", "Tipo", "Criado em"])


def listar_precos_banco() -> pd.DataFrame:
    inicializar_banco()
    with conectar_banco() as conn:
        rows = conn.execute(
            "SELECT sku, preco, origem, atualizado_em FROM precos ORDER BY sku"
        ).fetchall()
    return pd.DataFrame(rows, columns=["SKU", "Preço", "Origem", "Atualizado em"])


def listar_excluidos_banco() -> pd.DataFrame:
    inicializar_banco()
    with conectar_banco() as conn:
        rows = conn.execute(
            "SELECT sku, motivo, atualizado_em FROM skus_excluidos ORDER BY sku"
        ).fetchall()
    return pd.DataFrame(rows, columns=["SKU", "Motivo", "Atualizado em"])


# ============================================================
# Algoritmo de divisão operacional
# ============================================================
def ceil_decimal(a: Decimal, b: Decimal) -> int:
    return int((a / b).to_integral_value(rounding="ROUND_CEILING"))


def floor_decimal(a: Decimal, b: Decimal) -> int:
    return int((a / b).to_integral_value(rounding="ROUND_FLOOR"))


def totais_por_alocacao(alocacao: List[Dict[int, int]], itens: List[ItemPedido]) -> List[Decimal]:
    totais = []
    for pedido in alocacao:
        total = Decimal("0")
        for idx, qtd in pedido.items():
            total += itens[idx].preco_unitario * Decimal(qtd)
        totais.append(total.quantize(MOEDA_Q, rounding=ROUND_HALF_UP))
    return totais



def calcular_alvos_diversificados(total_geral: Decimal, minimo: Decimal, maximo: Decimal, n: int, permitir_residual: bool) -> List[Decimal]:
    """
    Monta alvos diferentes para cada pedido/NF.

    A ideia é evitar que a divisão inicial distribua todos os SKUs de forma
    proporcional e gere totais/quantidades praticamente iguais. Os alvos são
    usados somente como guia de composição; preço unitário e quantidade total
    continuam preservados.
    """
    if n <= 1:
        return [total_geral.quantize(MOEDA_Q, rounding=ROUND_HALF_UP)]

    amplitude_faixa = maximo - minimo
    centavo = Decimal("0.01")

    # Cenário residual: não há como todos ficarem dentro da faixa.
    # Mantém os primeiros pedidos em valores distintos dentro da faixa e deixa
    # o último como residual do pedido completo.
    if permitir_residual:
        alvos: List[Decimal] = []
        restante = total_geral
        for i in range(n - 1):
            pedidos_restantes = n - i - 1
            minimo_alvo = max(minimo, restante - (maximo * Decimal(pedidos_restantes)))
            maximo_alvo = min(maximo, restante - (centavo * Decimal(pedidos_restantes)))

            if maximo_alvo < minimo_alvo:
                alvo = max(centavo, min(maximo, maximo_alvo))
            else:
                if i % 3 == 0:
                    proporcao = Decimal("0.92")
                elif i % 3 == 1:
                    proporcao = Decimal("0.28")
                else:
                    proporcao = Decimal("0.63")
                alvo = minimo + (amplitude_faixa * proporcao)
                alvo = max(minimo_alvo, min(maximo_alvo, alvo))

            alvo = alvo.quantize(MOEDA_Q, rounding=ROUND_HALF_UP)
            alvos.append(alvo)
            restante -= alvo

        alvos.append(restante.quantize(MOEDA_Q, rounding=ROUND_HALF_UP))
        return alvos

    # Cenário matematicamente viável: cria variação simétrica ao redor da média,
    # mantendo a soma dos alvos igual ao total geral e respeitando a faixa.
    base = (total_geral / Decimal(n)).quantize(MOEDA_Q, rounding=ROUND_HALF_UP)
    folga_baixo = max(Decimal("0"), base - minimo)
    folga_cima = max(Decimal("0"), maximo - base)
    amplitude = min(amplitude_faixa * Decimal("0.35"), folga_baixo, folga_cima)

    if amplitude < Decimal("0.01"):
        amplitude = min(Decimal("25.00"), folga_baixo, folga_cima)

    if amplitude < Decimal("0.01"):
        return [base for _ in range(n - 1)] + [(total_geral - base * Decimal(n - 1)).quantize(MOEDA_Q, rounding=ROUND_HALF_UP)]

    pesos = []
    meio = Decimal(n - 1) / Decimal("2")
    denom = meio if meio != 0 else Decimal("1")
    for i in range(n):
        pesos.append((Decimal(i) - meio) / denom)

    # Alterna a ordem para os pedidos não ficarem com uma progressão tão regular.
    pesos_ordenados = []
    esquerda = 0
    direita = len(pesos) - 1
    while esquerda <= direita:
        pesos_ordenados.append(pesos[direita])
        if esquerda != direita:
            pesos_ordenados.append(pesos[esquerda])
        esquerda += 1
        direita -= 1

    alvos = [(base + amplitude * peso).quantize(MOEDA_Q, rounding=ROUND_HALF_UP) for peso in pesos_ordenados[:n]]

    # Ajuste fino de arredondamento no último alvo.
    diferenca = total_geral - sum(alvos, Decimal("0"))
    alvos[-1] = (alvos[-1] + diferenca).quantize(MOEDA_Q, rounding=ROUND_HALF_UP)

    # Garante limites após o ajuste fino.
    for i, alvo in enumerate(alvos):
        if alvo < minimo:
            falta = minimo - alvo
            alvos[i] = minimo
            for j in range(len(alvos)):
                if j != i and alvos[j] - falta >= minimo:
                    alvos[j] -= falta
                    break
        elif alvo > maximo:
            excesso = alvo - maximo
            alvos[i] = maximo
            for j in range(len(alvos)):
                if j != i and alvos[j] + excesso <= maximo:
                    alvos[j] += excesso
                    break

    return [a.quantize(MOEDA_Q, rounding=ROUND_HALF_UP) for a in alvos]


def alocar_por_alvos(itens: List[ItemPedido], alvos: List[Decimal], maximo: Decimal) -> Tuple[List[Dict[int, int]], List[Decimal]]:
    """Distribui unidades inteiras guiado por alvos diferentes de valor."""
    n = len(alvos)
    alocacao: List[Dict[int, int]] = [defaultdict(int) for _ in range(n)]
    totais = [Decimal("0") for _ in range(n)]

    itens_ordenados = sorted(itens, key=lambda x: (x.preco_unitario, x.valor_total, x.sku), reverse=True)

    for item in itens_ordenados:
        restante_qtd = item.quantidade
        preco = item.preco_unitario

        while restante_qtd > 0:
            candidatos = []
            for b in range(n):
                limite_bucket = min(alvos[b], maximo)
                espaco = limite_bucket - totais[b]
                if espaco >= preco:
                    candidatos.append((espaco, b))

            if candidatos:
                espaco, bucket = max(candidatos, key=lambda x: (x[0], -x[1]))
                qtd_add = int((espaco / preco).to_integral_value(rounding="ROUND_FLOOR"))
                qtd_add = max(1, min(restante_qtd, qtd_add))
            else:
                # Se nenhum alvo comporta mais uma unidade, usa o pedido com menor total
                # que ainda caiba no máximo. Isso preserva o pedido completo sem estourar NF.
                candidatos2 = [b for b in range(n) if totais[b] + preco <= maximo]
                if candidatos2:
                    bucket = min(candidatos2, key=lambda b: (totais[b], b))
                else:
                    bucket = min(range(n), key=lambda b: (totais[b], b))
                qtd_add = 1

            alocacao[bucket][item.indice] += qtd_add
            totais[bucket] += preco * Decimal(qtd_add)
            restante_qtd -= qtd_add

    return [dict(x) for x in alocacao], [t.quantize(MOEDA_Q, rounding=ROUND_HALF_UP) for t in totais]

def distribuir_itens(itens: List[ItemPedido], minimo: Decimal, maximo: Decimal) -> ResultadoDivisao:
    avisos: List[str] = []
    if not itens:
        return ResultadoDivisao(False, "Nenhum item válido para dividir.", [], [], avisos)
    if minimo <= 0 or maximo <= 0 or minimo > maximo:
        return ResultadoDivisao(False, "Informe uma faixa válida: mínimo maior que zero e menor ou igual ao máximo.", [], [], avisos)

    total_geral = sum((item.preco_unitario * Decimal(item.quantidade) for item in itens), Decimal("0")).quantize(MOEDA_Q, rounding=ROUND_HALF_UP)
    permitir_residual_fora_minimo = total_geral < minimo
    if total_geral < minimo:
        avisos.append(
            "O total geral é menor que o valor mínimo informado. O sistema gerará 1 pedido residual com o pedido completo."
        )

    item_unitario_maior = [item for item in itens if item.preco_unitario > maximo]
    if item_unitario_maior:
        skus = ", ".join(sorted({item.sku for item in item_unitario_maior})[:10])
        return ResultadoDivisao(False, f"Há SKU com preço unitário acima do máximo por pedido/NF: {skus}.", [], [], avisos)

    n_min = max(1, ceil_decimal(total_geral, maximo))
    n_max = floor_decimal(total_geral, minimo)
    if n_min > n_max:
        permitir_residual_fora_minimo = True
        n = n_min
        avisos.append(
            "O total geral não fecha matematicamente com todos os pedidos dentro da faixa. "
            "Para não travar a operação, o sistema gerará o pedido completo em NFs de melhor encaixe, "
            "mantendo todos os itens e podendo deixar pedido residual abaixo do mínimo."
        )
    else:
        alvo = (minimo + maximo) / Decimal("2")
        n_sugerido = int((total_geral / alvo).to_integral_value(rounding="ROUND_HALF_UP"))
        n = max(n_min, min(n_sugerido, n_max))
    n = max(1, n)

    # Alocação inicial guiada por alvos diferentes.
    # Isso evita NFs com valores e quantidades praticamente iguais, preservando
    # preço unitário, quantidade total e trilha de auditoria.
    alvos = calcular_alvos_diversificados(total_geral, minimo, maximo, n, permitir_residual_fora_minimo)
    alocacao, totais = alocar_por_alvos(itens, alvos, maximo)

    # Rebalanceamento por movimentação de unidades entre pedidos.
    # Objetivo: manter todos dentro da faixa sem alterar preço unitário nem quantidade total.
    max_iter = 50000
    for _ in range(max_iter):
        totais = totais_por_alocacao(alocacao, itens)
        baixos = [] if permitir_residual_fora_minimo else [i for i, t in enumerate(totais) if t < minimo]
        altos = [i for i, t in enumerate(totais) if t > maximo]
        if not baixos and not altos:
            msg = f"Divisão concluída em {n} pedido(s)/NF(s)."
            if permitir_residual_fora_minimo:
                msg = f"Divisão gerada em {n} pedido(s)/NF(s), preservando o pedido completo. Confira pedidos residuais abaixo do mínimo na auditoria."
            return ResultadoDivisao(True, msg, [dict(x) for x in alocacao], totais, avisos)

        movimento = False

        # Primeiro corrige pedidos abaixo do mínimo.
        if baixos:
            destino = min(baixos, key=lambda i: totais[i])
            necessidade = minimo - totais[destino]
            doadores = sorted([i for i in range(n) if i != destino and totais[i] > minimo], key=lambda i: totais[i], reverse=True)
            melhor = None
            melhor_score = None
            for origem in doadores:
                for idx, qtd_origem in alocacao[origem].items():
                    if qtd_origem <= 0:
                        continue
                    p = itens[idx].preco_unitario
                    if totais[destino] + p > maximo:
                        continue
                    if totais[origem] - p < minimo and totais[origem] <= maximo:
                        continue
                    # Aproxima o destino do mínimo sem estourar o máximo.
                    score = abs(necessidade - p)
                    if melhor is None or score < melhor_score:
                        melhor = (origem, destino, idx)
                        melhor_score = score
            if melhor:
                origem, destino, idx = melhor
                alocacao[origem][idx] -= 1
                if alocacao[origem][idx] == 0:
                    del alocacao[origem][idx]
                alocacao[destino][idx] += 1
                movimento = True

        if movimento:
            continue

        # Depois corrige pedidos acima do máximo.
        if altos:
            origem = max(altos, key=lambda i: totais[i])
            destinos = sorted([i for i in range(n) if i != origem and totais[i] < maximo], key=lambda i: totais[i])
            melhor = None
            melhor_score = None
            for destino in destinos:
                capacidade = maximo - totais[destino]
                for idx, qtd_origem in alocacao[origem].items():
                    if qtd_origem <= 0:
                        continue
                    p = itens[idx].preco_unitario
                    if p > capacidade:
                        continue
                    score = abs((totais[origem] - maximo) - p)
                    if melhor is None or score < melhor_score:
                        melhor = (origem, destino, idx)
                        melhor_score = score
            if melhor:
                origem, destino, idx = melhor
                alocacao[origem][idx] -= 1
                if alocacao[origem][idx] == 0:
                    del alocacao[origem][idx]
                alocacao[destino][idx] += 1
                movimento = True

        if not movimento:
            break

    totais = totais_por_alocacao(alocacao, itens)
    if permitir_residual_fora_minimo and all(t <= maximo for t in totais):
        return ResultadoDivisao(
            True,
            f"Divisão gerada em {n} pedido(s)/NF(s), preservando o pedido completo. Confira pedidos residuais abaixo do mínimo na auditoria.",
            [dict(x) for x in alocacao],
            totais,
            avisos,
        )

    fora = [fmt_moeda(t) for t in totais if not (minimo <= t <= maximo)]
    return ResultadoDivisao(
        False,
        "Não foi possível balancear todos os pedidos dentro da faixa usando quantidades inteiras. "
        f"Totais fora da faixa: {', '.join(fora[:8])}.",
        [dict(x) for x in alocacao],
        totais,
        avisos,
    )


# ============================================================
# Geração dos arquivos Excel
# ============================================================
def copiar_estilo(celula_origem, celula_destino) -> None:
    if celula_origem.has_style:
        celula_destino._style = copy(celula_origem._style)
    if celula_origem.number_format:
        celula_destino.number_format = celula_origem.number_format
    if celula_origem.alignment:
        celula_destino.alignment = copy(celula_origem.alignment)


def _normalizar_numero_base_pedido(base: object) -> str:
    """Normaliza o número original do pedido preservando zeros à esquerda quando houver."""
    if base in (None, ""):
        return "PEDIDO"
    texto = str(base).strip()
    # Quando o Excel entrega 37.0, transforma em 37.
    try:
        dec = Decimal(texto.replace(",", "."))
        if dec == dec.to_integral_value():
            return str(int(dec))
    except Exception:
        pass
    return texto


def numero_pedido_incrementado(base: object, seq: int) -> str:
    """
    Gera pedido sequencial a partir do número original.
    Ex.: base 37 e 3 pedidos -> 37, 38, 39.
    Ex.: base 00037 -> 00037, 00038, 00039.
    Para bases não numéricas, mantém a base e adiciona sufixo sequencial.
    """
    texto = _normalizar_numero_base_pedido(base)
    if texto.isdigit():
        largura = len(texto) if texto.startswith("0") else 0
        numero = int(texto) + seq - 1
        return str(numero).zfill(largura) if largura else str(numero)
    return f"{texto}-{seq:02d}"


def nome_pedido(base: str, seq: int, mascara: str) -> str:
    base_norm = _normalizar_numero_base_pedido(base)
    pedido_seq = numero_pedido_incrementado(base_norm, seq)
    mascara = (mascara or "{pedido_seq}").strip()
    try:
        return mascara.format(
            base=base_norm,
            seq=seq,
            seq2=f"{seq:02d}",
            seq3=f"{seq:03d}",
            pedido_seq=pedido_seq,
            pedido=pedido_seq,
        )
    except Exception:
        return pedido_seq


def gerar_planilha_pedido(
    pedido_bytes: bytes,
    aba: str,
    header_row: int,
    colmap: Dict[str, int],
    itens: List[ItemPedido],
    alocacao_pedido: Dict[int, int],
    total_pedido: Decimal,
    numero_pedido_saida: str,
) -> bytes:
    wb = load_workbook(BytesIO(pedido_bytes), data_only=False)
    ws = wb[aba]
    max_col = ws.max_column
    linha_modelo = header_row + 1

    estilo_modelo = []
    for col in range(1, max_col + 1):
        estilo_modelo.append(copy(ws.cell(linha_modelo, col)._style))
    altura_modelo = ws.row_dimensions[linha_modelo].height

    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)

    linha_destino = header_row + 1
    for idx in sorted(alocacao_pedido.keys(), key=lambda i: itens[i].linha_excel):
        qtd = alocacao_pedido[idx]
        if qtd <= 0:
            continue
        item = itens[idx]
        valores = list(item.valores_linha[:max_col])
        while len(valores) < max_col:
            valores.append(None)

        valores[colmap["numero_pedido"] - 1] = numero_pedido_saida
        valores[colmap["sku"] - 1] = item.sku
        valores[colmap["quantidade"] - 1] = qtd
        if colmap.get("valor_unitario"):
            valores[colmap["valor_unitario"] - 1] = float(item.preco_unitario)
        if colmap.get("valor_total"):
            valores[colmap["valor_total"] - 1] = float((item.preco_unitario * Decimal(qtd)).quantize(MOEDA_Q, rounding=ROUND_HALF_UP))
        if colmap.get("total_pedido"):
            valores[colmap["total_pedido"] - 1] = float(total_pedido.quantize(MOEDA_Q, rounding=ROUND_HALF_UP))

        for col, valor in enumerate(valores, start=1):
            cell = ws.cell(linha_destino, col)
            cell.value = valor
            if col - 1 < len(estilo_modelo):
                cell._style = copy(estilo_modelo[col - 1])
        if altura_modelo:
            ws.row_dimensions[linha_destino].height = altura_modelo
        linha_destino += 1

    # Formatos de número relevantes.
    if colmap.get("valor_unitario"):
        col = get_column_letter(colmap["valor_unitario"])
        for row in range(header_row + 1, linha_destino):
            ws[f"{col}{row}"].number_format = 'R$ #,##0.0000'
    if colmap.get("valor_total"):
        col = get_column_letter(colmap["valor_total"])
        for row in range(header_row + 1, linha_destino):
            ws[f"{col}{row}"].number_format = 'R$ #,##0.00'
    if colmap.get("total_pedido"):
        col = get_column_letter(colmap["total_pedido"])
        for row in range(header_row + 1, linha_destino):
            ws[f"{col}{row}"].number_format = 'R$ #,##0.00'
    if colmap.get("quantidade"):
        col = get_column_letter(colmap["quantidade"])
        for row in range(header_row + 1, linha_destino):
            ws[f"{col}{row}"].number_format = '0'

    saida = BytesIO()
    wb.save(saida)
    return saida.getvalue()


def gerar_auditoria(
    itens: List[ItemPedido],
    alocacao: List[Dict[int, int]],
    totais: List[Decimal],
    minimo: Decimal,
    maximo: Decimal,
    numeros_pedido: List[str],
    pendencias: Optional[List[Dict[str, object]]] = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo_NFs"
    ws.append(["Pedido/NF", "Valor total", "Dentro da faixa", "Qtd. linhas", "Qtd. peças"])
    for i, pedido in enumerate(alocacao):
        qtd_pecas = sum(pedido.values())
        dentro = "Sim" if minimo <= totais[i] <= maximo else "Não"
        ws.append([numeros_pedido[i], float(totais[i]), dentro, len(pedido), qtd_pecas])
    ws.append([])
    ws.append(["Total geral", float(sum(totais, Decimal("0"))), "", "", sum(sum(p.values()) for p in alocacao)])

    ws2 = wb.create_sheet("Itens")
    ws2.append(["Pedido/NF", "Linha original", "SKU", "Quantidade", "Preço unitário", "Valor total item"])
    for i, pedido in enumerate(alocacao):
        for idx in sorted(pedido.keys(), key=lambda x: itens[x].linha_excel):
            qtd = pedido[idx]
            item = itens[idx]
            ws2.append([
                numeros_pedido[i],
                item.linha_excel,
                item.sku,
                qtd,
                float(item.preco_unitario),
                float((item.preco_unitario * Decimal(qtd)).quantize(MOEDA_Q, rounding=ROUND_HALF_UP)),
            ])

    if pendencias:
        ws3 = wb.create_sheet("Pendencias")
        ws3.append(["Linha", "SKU", "Problema"])
        for p in pendencias:
            ws3.append([p.get("Linha"), p.get("SKU"), p.get("Problema")])

    for wsx in wb.worksheets:
        for row in wsx.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")
        for cell in wsx[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for col in range(1, wsx.max_column + 1):
            letra = get_column_letter(col)
            largura = max(12, min(38, max(len(str(wsx.cell(row, col).value or "")) for row in range(1, wsx.max_row + 1)) + 2))
            wsx.column_dimensions[letra].width = largura
        for row in range(2, wsx.max_row + 1):
            for col in range(1, wsx.max_column + 1):
                header = str(wsx.cell(1, col).value or "").lower()
                if "valor" in header or "preço" in header:
                    wsx.cell(row, col).number_format = 'R$ #,##0.00'
        wsx.freeze_panes = "A2"

    saida = BytesIO()
    wb.save(saida)
    return saida.getvalue()


def gerar_pendencias_xlsx(pendencias: List[Dict[str, object]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pendencias"
    ws.append(["Linha", "SKU", "Problema"])
    for p in pendencias:
        ws.append([p.get("Linha"), p.get("SKU"), p.get("Problema")])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C00000")
    for col in range(1, ws.max_column + 1):
        letra = get_column_letter(col)
        ws.column_dimensions[letra].width = 22 if col < 3 else 70
    saida = BytesIO()
    wb.save(saida)
    return saida.getvalue()


def montar_zip(
    pedido_bytes: bytes,
    aba_pedido: str,
    header_row: int,
    colmap: Dict[str, int],
    itens: List[ItemPedido],
    resultado: ResultadoDivisao,
    base_numero_pedido: str,
    mascara: str,
    minimo: Decimal,
    maximo: Decimal,
) -> bytes:
    zip_buffer = BytesIO()
    numeros = [nome_pedido(base_numero_pedido, i + 1, mascara) for i in range(len(resultado.alocacao))]
    with ZipFile(zip_buffer, "w", compression=ZIP_DEFLATED) as zf:
        for i, aloc in enumerate(resultado.alocacao):
            arquivo = gerar_planilha_pedido(
                pedido_bytes=pedido_bytes,
                aba=aba_pedido,
                header_row=header_row,
                colmap=colmap,
                itens=itens,
                alocacao_pedido=aloc,
                total_pedido=resultado.totais[i],
                numero_pedido_saida=numeros[i],
            )
            nome_arquivo = f"pedido_{i + 1:02d}_{numeros[i]}.xlsx".replace("/", "-").replace("\\", "-")
            zf.writestr(nome_arquivo, arquivo)

        auditoria = gerar_auditoria(
            itens=itens,
            alocacao=resultado.alocacao,
            totais=resultado.totais,
            minimo=minimo,
            maximo=maximo,
            numeros_pedido=numeros,
        )
        zf.writestr("AUDITORIA_desmembramento.xlsx", auditoria)
    return zip_buffer.getvalue()


# ============================================================
# Interface Streamlit
# ============================================================
st.set_page_config(page_title="Gerador de pedidos desmembrados", layout="wide")
inicializar_banco()

st.title("Gerador de pedidos desmembrados por faixa de valor")
st.caption("Mantém preço unitário, quantidade total e layout do modelo. Gera uma trilha de auditoria para conferência.")

with st.expander("Critérios usados pelo sistema", expanded=False):
    st.write(
        "O sistema divide o pedido por composição de itens, sem alterar preço unitário e sem alterar a quantidade total. "
        "A distribuição usa alvos diferentes por pedido/NF para evitar totais e quantidades iguais. "
        "Quando o total geral não permitir que todos os pedidos fiquem dentro da faixa, o sistema mantém o pedido completo e sinaliza o residual na auditoria."
    )

with st.expander("Banco de preços salvo no sistema", expanded=False):
    st.caption("As listas importadas e os preços manuais ficam salvos no banco SQLite local do app.")
    col_b1, col_b2, col_b3 = st.columns(3)
    with col_b1:
        st.write("**Fontes importadas**")
        st.dataframe(listar_fontes_salvas(), use_container_width=True, hide_index=True)
    with col_b2:
        st.write("**Preços cadastrados**")
        df_pre_banco = listar_precos_banco()
        st.dataframe(df_pre_banco.tail(200), use_container_width=True, hide_index=True)
    with col_b3:
        st.write("**SKUs excluídos**")
        st.dataframe(listar_excluidos_banco(), use_container_width=True, hide_index=True)

col1, col2 = st.columns(2)
with col1:
    pedido_file = st.file_uploader("Planilha de pedido padronizada (.xlsx)", type=["xlsx"], key="upload_pedido")
with col2:
    preco_file = st.file_uploader("Lista de preços (.xlsx ou PDF do Bling)", type=["xlsx", "pdf"], key="upload_preco")

# Mantém os arquivos e escolhas em memória da sessão.
# Isso evita que, após clicar em "Salvar ajustes e reprocessar", o Streamlit volte para a tela inicial.
if pedido_file is not None:
    st.session_state["pedido_bytes_mem"] = pedido_file.getvalue()
    st.session_state["pedido_nome_mem"] = pedido_file.name

if preco_file is not None:
    st.session_state["preco_bytes_mem"] = preco_file.getvalue()
    st.session_state["preco_nome_mem"] = preco_file.name

pedido_bytes_mem = st.session_state.get("pedido_bytes_mem")
preco_bytes_mem = st.session_state.get("preco_bytes_mem")
pedido_nome_mem = st.session_state.get("pedido_nome_mem", "pedido.xlsx")
preco_nome_mem = st.session_state.get("preco_nome_mem", "lista_precos")

tem_pedido = bool(pedido_bytes_mem)
tem_preco = bool(preco_bytes_mem)

if tem_pedido and pedido_file is None:
    st.info(f"Pedido carregado na sessão: {pedido_nome_mem}")
if tem_preco and preco_file is None:
    st.info(f"Lista de preços carregada na sessão: {preco_nome_mem}")

min_col, max_col, suf_col = st.columns([1, 1, 2])
with min_col:
    valor_min = st.number_input("Valor mínimo por pedido/NF", min_value=0.01, value=110000.00, step=1000.00, format="%.2f")
with max_col:
    valor_max = st.number_input("Valor máximo por pedido/NF", min_value=0.01, value=125000.00, step=1000.00, format="%.2f")
with suf_col:
    mascara_pedido = st.text_input(
        "Máscara do número do pedido de saída",
        value="{pedido_seq}",
        help="Use {pedido_seq} para seguir a numeração original. Ex.: pedido 37 com 3 saídas gera 37, 38 e 39. Também aceita {base}, {seq}, {seq2} e {seq3}.",
    )

aba_pedido = st.session_state.get("aba_pedido_mem")
aba_preco = st.session_state.get("aba_preco_mem")
nome_lista = st.session_state.get("nome_lista_mem")

if tem_pedido:
    wb_pedido_ro = load_workbook(BytesIO(pedido_bytes_mem), read_only=True, data_only=True)
    sheetnames = wb_pedido_ro.sheetnames
    idx_aba = sheetnames.index(aba_pedido) if aba_pedido in sheetnames else 0
    aba_pedido = st.selectbox("Aba do pedido", sheetnames, index=idx_aba)
    st.session_state["aba_pedido_mem"] = aba_pedido

if tem_preco:
    tipo_preco = identificar_tipo_preco(preco_nome_mem)
    if tipo_preco == "xlsx":
        abas_preco, nomes_listas = listar_abas_e_listas_preco(preco_bytes_mem)
        idx_aba_preco = abas_preco.index(aba_preco) if aba_preco in abas_preco else 0
        aba_preco = st.selectbox("Aba da lista de preços", abas_preco, index=idx_aba_preco)
        st.session_state["aba_preco_mem"] = aba_preco
        if nomes_listas:
            opcoes = ["Todas as listas"] + nomes_listas
            escolha_atual = nome_lista if nome_lista in nomes_listas else (opcoes[1] if len(opcoes) > 1 else opcoes[0])
            idx_lista = opcoes.index(escolha_atual) if escolha_atual in opcoes else 0
            escolha = st.selectbox("Nome da lista de preços", opcoes, index=idx_lista)
            nome_lista = None if escolha == "Todas as listas" else escolha
            st.session_state["nome_lista_mem"] = nome_lista
        else:
            nome_lista = None
            st.session_state["nome_lista_mem"] = None
    else:
        aba_preco = "PDF Bling"
        nome_lista = None
        st.session_state["aba_preco_mem"] = aba_preco
        st.session_state["nome_lista_mem"] = None
        st.info("PDF identificado. O sistema usará a coluna 'R$ Preço da lista' do relatório do Bling.")

processar = st.button("Gerar pedidos", type="primary", disabled=not (tem_pedido and tem_preco and aba_pedido and aba_preco))
if processar:
    st.session_state["processamento_ativo"] = True

executar_geracao = bool(st.session_state.get("processamento_ativo", False)) or bool(st.session_state.pop("reprocessar_apos_ajuste", False))

if executar_geracao:
    try:
        pedido_bytes = st.session_state.get("pedido_bytes_mem")
        preco_bytes = st.session_state.get("preco_bytes_mem")
        if not pedido_bytes or not preco_bytes:
            st.session_state["processamento_ativo"] = False
            st.warning("Reenvie o pedido e a lista de preços para continuar.")
            st.stop()
        minimo = para_decimal(valor_min, "valor mínimo").quantize(MOEDA_Q, rounding=ROUND_HALF_UP)
        maximo = para_decimal(valor_max, "valor máximo").quantize(MOEDA_Q, rounding=ROUND_HALF_UP)

        tipo_preco_exec = identificar_tipo_preco(preco_nome_mem)
        if tipo_preco_exec == "pdf":
            precos_upload = carregar_precos_pdf_bling(preco_bytes)
        else:
            precos_upload = carregar_precos(preco_bytes, aba_preco, nome_lista)

        salvar_fonte_preco(preco_nome_mem, tipo_preco_exec, preco_bytes)
        salvar_precos_no_banco(precos_upload, f"upload:{preco_nome_mem}")

        precos = carregar_precos_do_banco()
        precos.update(precos_upload)
        skus_excluidos = carregar_skus_excluidos()
        header_row, colmap, itens, pendencias = carregar_itens_pedido(pedido_bytes, aba_pedido, precos, skus_excluidos)

        st.subheader("Conferência inicial")
        total_lido = sum((i.valor_total for i in itens), Decimal("0")).quantize(MOEDA_Q, rounding=ROUND_HALF_UP)
        c1, c2, c3 = st.columns(3)
        c1.metric("Itens válidos", len(itens))
        c2.metric("Total precificado", fmt_moeda(total_lido))
        c3.metric("Pendências", len(pendencias))

        if pendencias:
            st.error("Existem pendências antes da divisão. Inclua o preço manual ou marque para excluir o SKU do pedido.")
            df_pend = pd.DataFrame(pendencias)
            df_editor = df_pend.copy()
            df_editor["Preço manual"] = ""
            df_editor["Excluir do pedido"] = False
            editado = st.data_editor(
                df_editor,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Preço manual": st.column_config.TextColumn(
                        "Preço manual",
                        help="Informe no padrão brasileiro, exemplo: 12,63",
                    ),
                    "Excluir do pedido": st.column_config.CheckboxColumn(
                        "Excluir do pedido",
                        help="Marca o SKU para ser ignorado nas próximas gerações.",
                    ),
                },
                disabled=["Linha", "SKU", "Problema"],
                key="editor_pendencias",
            )
            col_p1, col_p2 = st.columns([1, 3])
            with col_p1:
                if st.button("Salvar ajustes e reprocessar", type="primary"):
                    try:
                        qtd_salva = salvar_precos_manuais(editado.to_dict("records"))
                        st.success(f"{qtd_salva} ajuste(s) salvo(s). Reprocessando automaticamente...")
                        st.session_state["reprocessar_apos_ajuste"] = True
                        st.rerun()
                    except Exception as exc:
                        st.exception(exc)
            with col_p2:
                st.download_button(
                    "Baixar pendências em Excel",
                    data=gerar_pendencias_xlsx(pendencias),
                    file_name="pendencias_skus.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            st.stop()

        resultado = distribuir_itens(itens, minimo, maximo)
        if not resultado.sucesso:
            st.error(resultado.mensagem)
            if resultado.totais:
                st.dataframe(
                    pd.DataFrame({
                        "Pedido/NF": [i + 1 for i in range(len(resultado.totais))],
                        "Total": [fmt_moeda(t) for t in resultado.totais],
                        "Dentro da faixa": ["Sim" if minimo <= t <= maximo else "Não" for t in resultado.totais],
                    }),
                    use_container_width=True,
                )
            st.stop()

        # Número base do pedido original.
        col_num = colmap["numero_pedido"]
        wb_tmp = load_workbook(BytesIO(pedido_bytes), data_only=True, read_only=True)
        ws_tmp = wb_tmp[aba_pedido]
        base_numero = ws_tmp.cell(header_row + 1, col_num).value or "PEDIDO"

        numeros = [nome_pedido(str(base_numero), i + 1, mascara_pedido) for i in range(len(resultado.alocacao))]
        resumo = pd.DataFrame({
            "Pedido/NF": numeros,
            "Valor total": [fmt_moeda(t) for t in resultado.totais],
            "Qtd. linhas": [len(a) for a in resultado.alocacao],
            "Qtd. peças": [sum(a.values()) for a in resultado.alocacao],
            "Dentro da faixa": ["Sim" if minimo <= t <= maximo else "Não" for t in resultado.totais],
        })

        st.success(resultado.mensagem)
        for aviso in resultado.avisos:
            st.warning(aviso)
        st.dataframe(resumo, use_container_width=True)

        zip_bytes = montar_zip(
            pedido_bytes=pedido_bytes,
            aba_pedido=aba_pedido,
            header_row=header_row,
            colmap=colmap,
            itens=itens,
            resultado=resultado,
            base_numero_pedido=str(base_numero),
            mascara=mascara_pedido,
            minimo=minimo,
            maximo=maximo,
        )
        st.download_button(
            "Baixar ZIP com pedidos e auditoria",
            data=zip_bytes,
            file_name="pedidos_desmembrados.zip",
            mime="application/zip",
        )

    except Exception as exc:
        st.exception(exc)
